#-*- encoding: utf-8 -*-
'''
Created on 2016年8月25日

@author: BirdLin
'''

from OCVTable import Cell_OCV_Table
import openpyxl
C_mAh = 0
C_Wh = 0
Current_Status =''

#--- Constant Variables -----------------------------------------------
Cell_Series = 10                        # number
Cell_Parallel = 1                       # number
Cell_Capacity = 2600                    # mAh
Design_Capacity = Cell_Parallel * Cell_Capacity
Full_SOC = 100                          # %
Charge_Update_Vol = 39000

Discharge_Current_Threshold = -190       # mA
Charge_Current_Threshold = 100          # mA
Charge_Update_Current_mA = 200          # mA
Charge_Update_Time = 100                 # sec
Static_Update_Time = 800               # sec
Static_FCCUpdate_Percentage = 30        # %
Long_Static_Update_Time = 1600        # sec
avg_Vol_num = 5
Dynamic_Voltage_Threshold = 50          # mV
RM_percentage_display_buffer = 5        # %

#--- Tracking & Initial Variables -------------------------------------

Full_Charge_Capacity_mAh = Cell_Parallel * Cell_Capacity                # mAh

#RM_percentage = (RM_mAh * 100.0 / Full_Charge_Capacity_mAh)
#Used_Capacity_mAh = Full_Charge_Capacity_mAh - RM_mAh ### The capacity has been used.

Discharged_Capacity_calculation = 0           ### Discharged capacity by count minute
Charged_Capacity_calculation = 0              ### Charged capacity by count minute
Extra_Dsg_Cap = 0                               ### Extra discharged capacity

Time_Interval = 0
Rest_Time = 0                                    # sec
Rest_Time_Last = 0
Rest_Time_for_Update = 0                        # sec
Charge_Time = 0                                 # sec
Charge_Tape_Time = 0                            # sec
Discharge_Time = 0                              # sec
RM_percentage_display_old = 0                   # %
RM_percentage_display_middle = 0                   # %

#Static_Update_Enable = False                   # False means can't update or just update; True means available to update.
CurrentInCHG = False
CurrentInDSG = False
CurrentInSTATIC = True
Charge_FCC_is_Updated = False                 # False is available to update; True means just updated now.
Initialized_Flag = True                        # True means it's in initialized; False means already initialized
Vol_is_Dynamic = True

### To get static SOC%
def Get_Static_OCV_Percentage():
    i = 0
    global Total_Vol, RM_percentage, Cell_OCV_Table, Cell_Series
    if Total_Vol >= Cell_OCV_Table[0] * Cell_Series :
        RM_percentage = Full_SOC
    elif Total_Vol <= Cell_OCV_Table[99] * Cell_Series :
        RM_percentage = 0
    else :
        while Total_Vol < Cell_OCV_Table[i] * Cell_Series :
            i += 1
            RM_percentage = Full_SOC - i

### To get static capacity
def Normal_Capacity_Calculation():
    global RM_mAh, RM_percentage, Used_Capacity_mAh, Discharged_Capacity_calculation, Charged_Capacity_calculation, RM_percentage
    Used_Capacity_mAh = (100 - RM_percentage)/100 * Full_Charge_Capacity_mAh
    RM_mAh = Full_Charge_Capacity_mAh - Used_Capacity_mAh
    Discharged_Capacity_calculation = 0
    Charged_Capacity_calculation = 0

def Initialization():
    global Rest_Time, Rest_Time_for_Update, Initialized_Flag, RM_percentage_display_old, RM_percentage_display_middle, RM_percentage_display, RM_percentage
    if((Rest_Time >= Long_Static_Update_Time and Vol_is_Dynamic == False) or Initialized_Flag == True):
        Get_Static_OCV_Percentage()
        Normal_Capacity_Calculation()
        Initialized_Flag = False
        Rest_Time_for_Update = 0
        Rest_Time = 0
        RM_percentage_display_old = RM_percentage_display_middle = RM_percentage_display = RM_percentage

def Check_Current_Status():
    global Current_mA, CurrentInCHG, CurrentInDSG, CurrentInSTATIC, Current_Status
    if(Current_mA >= Charge_Current_Threshold):
        CurrentInCHG = True
        CurrentInDSG = False
        CurrentInSTATIC = False
        Current_Status = 'Charging'
    elif(Current_mA <= Discharge_Current_Threshold):
        CurrentInCHG = False
        CurrentInDSG = True
        CurrentInSTATIC = False
        Current_Status = 'Discharging'
    else:
        CurrentInCHG = False
        CurrentInDSG = False
        CurrentInSTATIC = True
        Current_Status = 'Rest'

def Calculation_Charging():
    global Current_mA, Charge_Tape_Time, Time_Interval, Rest_Time, Rest_Time_for_Update, RM_mAh, RM_percentage, Used_Capacity_mAh, Discharged_Capacity_calculation, Charged_Capacity_calculation, Full_Charge_Capacity_mAh, Extra_Dsg_Cap, Charge_FCC_is_Updated
    Rest_Time = 0
    Rest_Time_for_Update = 0
    Charged_Capacity_calculation += abs(Current_mA * Time_Interval/3600)
    Used_Capacity_mAh += Discharged_Capacity_calculation         ## UC update by discharge capacity
    Used_Capacity_mAh -= Charged_Capacity_calculation            ## UC update by charge capacity
    RM_mAh = Full_Charge_Capacity_mAh - Used_Capacity_mAh
    if(RM_mAh >= Full_Charge_Capacity_mAh):                      ## RM is more than full
        Full_Charge_Capacity_mAh = RM_mAh                        ## FCC update when charging
        RM_percentage = 100
        Used_Capacity_mAh = 0
    elif(RM_mAh < 0):
        RM_mAh = 0
        RM_percentage = 0
    else:                                                        ## RM_mAh >= 0
        RM_percentage = RM_mAh * 100.0 / Full_Charge_Capacity_mAh 
    Charged_Capacity_calculation = 0
    Discharged_Capacity_calculation = 0
    if(Current_mA <= Charge_Update_Current_mA):        ## FCC Charge Update Criteria
        Charge_Tape_Time += Time_Interval
        if(Charge_Tape_Time >= Charge_Update_Time and RM_mAh <= Full_Charge_Capacity_mAh and Total_Vol >= Charge_Update_Vol and Charge_FCC_is_Updated == False):
            Used_Capacity_mAh = 0
            Full_Charge_Capacity_mAh = RM_mAh
            RM_percentage = 100
            Charge_FCC_is_Updated = True
            Charge_Tape_Time = 0
    Extra_Dsg_Cap = 0                                   ## Only update during discharging and rest

def Calculation_Discharging():
    global Current_mA, Rest_Time, Time_Interval, Charge_Tape_Time, Rest_Time_for_Update, RM_mAh, RM_percentage, Used_Capacity_mAh, Discharged_Capacity_calculation, Charged_Capacity_calculation, Full_Charge_Capacity_mAh, Extra_Dsg_Cap, Charge_FCC_is_Updated
    Rest_Time = 0
    Rest_Time_for_Update = 0
    Charge_Tape_Time = 0
    Discharged_Capacity_calculation += abs(Current_mA * Time_Interval/3600)
    Used_Capacity_mAh += Discharged_Capacity_calculation         ## UC update by discharge capacity
    Used_Capacity_mAh -= Charged_Capacity_calculation            ## UC update by charge capacity
    RM_mAh = Full_Charge_Capacity_mAh - Used_Capacity_mAh
    if(RM_mAh >= Full_Charge_Capacity_mAh):
        RM_mAh = Full_Charge_Capacity_mAh                        ## FCC is max RM_mAh when discharging
        RM_percentage = 100
    elif(RM_mAh < 0):
        Used_Capacity_mAh = Full_Charge_Capacity_mAh
        Extra_Dsg_Cap += Discharged_Capacity_calculation         ## count the extra discharge capacity
        RM_mAh = 0
        RM_percentage = 0
    else:                                                        ## RM_mAh >= 0
        RM_percentage = RM_mAh * 100.0 / Full_Charge_Capacity_mAh 
    Charged_Capacity_calculation = 0
    Discharged_Capacity_calculation = 0
    if(RM_percentage <= 70):
        Charge_FCC_is_Updated = False

def Calculation_Rest():
    global Rest_Time, Rest_Time_for_Update, Time_Interval, Charge_Tape_Time, Total_Vol, RM_mAh, RM_percentage, Used_Capacity_mAh, Discharged_Capacity_calculation, Charged_Capacity_calculation, Full_Charge_Capacity_mAh, Extra_Dsg_Cap, Initialized_Flag
    Initialization()
    Charge_Tape_Time = 0                                                ## Need to be adjusted if it's not started in rest
    Rest_Time += Time_Interval
    Rest_Time_for_Update += Time_Interval                                             ## count rest time
    ## Static Update
    if(Rest_Time_for_Update >= Static_Update_Time and Vol_is_Dynamic == False):
        if(RM_percentage < Static_FCCUpdate_Percentage):
            Get_Static_OCV_Percentage()
            Rest_Time_for_Update = 0
            Used_Capacity_mAh += Extra_Dsg_Cap
            Full_Charge_Capacity_mAh = ((Used_Capacity_mAh*100)/(100-RM_percentage))
            RM_mAh = Full_Charge_Capacity_mAh - Used_Capacity_mAh
            RM_percentage = RM_mAh * 100.0 / (Full_Charge_Capacity_mAh +1)
            Charged_Capacity_calculation = 0
            Discharged_Capacity_calculation = 0
            Extra_Dsg_Cap = 0
    elif(Rest_Time_for_Update < Static_Update_Time):   
        Used_Capacity_mAh -= Charged_Capacity_calculation            ## UC update by charge capacity
        Used_Capacity_mAh += Discharged_Capacity_calculation         ## UC update by discharge capacity
        RM_mAh = Full_Charge_Capacity_mAh - Used_Capacity_mAh
        if(RM_mAh >= Full_Charge_Capacity_mAh):
            Full_Charge_Capacity_mAh = RM_mAh                        ## FCC update when Charging
            RM_percentage = 100
        elif(RM_mAh < 0):
            Used_Capacity_mAh = Full_Charge_Capacity_mAh
            Extra_Dsg_Cap += Discharged_Capacity_calculation         ## count the extra discharge capacity
            RM_mAh = 0
            RM_percentage = 0
        else:                                                        ## RM_mAh >= 0
            RM_percentage = RM_mAh * 100.0 / Full_Charge_Capacity_mAh 
        Charged_Capacity_calculation = 0
        Discharged_Capacity_calculation = 0

def AverageVol():
    global vol_List, vol_List_index, j, avg_Vol
    if vol_List_index >= avg_Vol_num :
        vol_List_index = avg_Vol_num
    j = vol_List_index-1
    while j > 0 :
        vol_List[j] = vol_List[j-1]
        j -= 1
    vol_List[0] = Total_Vol
    avg_Vol = sum(vol_List)/vol_List_index
    vol_List_index+=1

def RM_Percentage_Display_Calculation():
    global RM_percentage, RM_percentage_display
    if(RM_percentage >= 100):
        RM_percentage_display = 100
    else:
        RM_percentage_display = (100-(100-RM_percentage)*100/(100-RM_percentage_display_buffer))
    if(RM_percentage_display <= 0):
        RM_percentage_display = 0
    return RM_percentage_display

def RM_Percentage_Display_smooth():
    global RM_percentage_display_old, RM_percentage_display_middle, RM_percentage_display
    RM_percentage_display_old = RM_percentage_display_middle
    RM_percentage_display_middle = RM_percentage_display
    if((RM_percentage_display - RM_percentage_display_old) > 1.5):
        RM_percentage_display = RM_percentage_display_old + 0.2
        RM_percentage_display_middle = RM_percentage_display
    elif((RM_percentage_display - RM_percentage_display_old) < -1.5):
        RM_percentage_display = RM_percentage_display_old - 0.2
        RM_percentage_display_middle = RM_percentage_display

###### Main Gauge Flow ######
#### Reading Data
wb = openpyxl.load_workbook('rest+discharge.xlsx', read_only=False)
ws = wb.get_sheet_by_name ('sheet1')

#### Initialization
## Initialized in rest condition

#### Input Data to Algorithm
row_index = 1
avg_Vol = 0
vol_List = [0] * avg_Vol_num
vol_List_index = 1
print ('Calculating......')

for row in ws.rows:
    if row_index == 1 :
        ws.cell(row=row_index, column=len(row)+1).value='C_mAh'
        ws.cell(row=row_index, column=len(row)+2).value='C_Wh'
        ws.cell(row=row_index, column=len(row)+3).value='mV'
        ws.cell(row=row_index, column=len(row)+4).value='mV_avg'
        ws.cell(row=row_index, column=len(row)+5).value='mA'
        ws.cell(row=row_index, column=len(row)+6).value='Current_Status'
        ws.cell(row=row_index, column=len(row)+7).value= 'Rest_Time'
        ws.cell(row=row_index, column=len(row)+8).value= 'Rest_Time_for_Update'
        ws.cell(row=row_index, column=len(row)+9).value= 'Charge_Tape_Time'
        ws.cell(row=row_index, column=len(row)+10).value= 'Charged_Capacity_calculation'
        ws.cell(row=row_index, column=len(row)+11).value= 'Discharged_Capacity_calculation'
        ws.cell(row=row_index, column=len(row)+12).value= 'Extra_Dsg_Cap'
        ws.cell(row=row_index, column=len(row)+13).value= 'Full_Charge_Capacity_mAh'
        ws.cell(row=row_index, column=len(row)+14).value= 'RM_mAh'
        ws.cell(row=row_index, column=len(row)+15).value= 'RM_percentage'
        ws.cell(row=row_index, column=len(row)+16).value= 'RM_percentage_display'
        ws.cell(row=row_index, column=len(row)+17).value= 'Used_Capacity_mAh'
        ws.cell(row=row_index, column=len(row)+18).value= 'Charge_FCC_is_Updated'
        ws.cell(row=row_index, column=len(row)+19).value= 'Vol_is_Dynamic'
        row_index += 1
        continue
    Total_Vol = row[0].value * 1000
    Current_mA = row[1].value * 1000
    Time_Interval = row[2].value
    C_mAh += Current_mA * (-Time_Interval) /3600
    C_Wh += Current_mA/-1000 * Total_Vol/1000 * Time_Interval /3600
#### Current Status Check
    Check_Current_Status()

## Average Voltage
    AverageVol()
    if(abs(Total_Vol-avg_Vol) <= Dynamic_Voltage_Threshold and CurrentInSTATIC == True):
        Vol_is_Dynamic = False
    else:
        Vol_is_Dynamic = True

#### Capacity Calculaton
## Charging                        
    if(CurrentInCHG == True):
        Calculation_Charging()
## Discharging
    elif(CurrentInDSG == True):
        Calculation_Discharging()
## Rest
    elif(CurrentInSTATIC == True):
        Calculation_Rest()

## Display RM Percentage Calculation
    RM_Percentage_Display_Calculation()
    RM_Percentage_Display_smooth()
   
    ws.cell(row=row_index, column=len(row)+1).value = C_mAh
    ws.cell(row=row_index, column=len(row)+2).value = C_Wh    
    ws.cell(row=row_index, column=len(row)+3).value = Total_Vol
    ws.cell(row=row_index, column=len(row)+4).value = avg_Vol
    ws.cell(row=row_index, column=len(row)+5).value = Current_mA
    ws.cell(row=row_index, column=len(row)+6).value = Current_Status
    ws.cell(row=row_index, column=len(row)+7).value= Rest_Time
    ws.cell(row=row_index, column=len(row)+8).value= Rest_Time_for_Update
    ws.cell(row=row_index, column=len(row)+9).value= Charge_Tape_Time
    ws.cell(row=row_index, column=len(row)+10).value= Charged_Capacity_calculation
    ws.cell(row=row_index, column=len(row)+11).value= Discharged_Capacity_calculation
    ws.cell(row=row_index, column=len(row)+12).value= Extra_Dsg_Cap
    ws.cell(row=row_index, column=len(row)+13).value= Full_Charge_Capacity_mAh
    ws.cell(row=row_index, column=len(row)+14).value= RM_mAh
    ws.cell(row=row_index, column=len(row)+15).value= RM_percentage
    ws.cell(row=row_index, column=len(row)+16).value= RM_percentage_display
    ws.cell(row=row_index, column=len(row)+17).value= Used_Capacity_mAh
    ws.cell(row=row_index, column=len(row)+18).value= Charge_FCC_is_Updated
    ws.cell(row=row_index, column=len(row)+19).value= Vol_is_Dynamic

    row_index += 1

#### Save data
wb.save(filename = 'new_rest+discharge.xlsx')
print ('Calculation Finished !')
